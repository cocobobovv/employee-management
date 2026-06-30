# -*- coding: utf-8 -*-
"""
员工管理系统 — Flask 主应用
功能：CRUD、搜索筛选、数据导出、RESTful API
"""
import csv
import io
import os
import re
from datetime import date, datetime
from typing import Optional

from flask import (Flask, render_template, request, redirect,
                   url_for, flash, jsonify, send_file, Response, session)
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook, load_workbook

from models import db, Employee, User
from sqlalchemy.exc import SQLAlchemyError

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY',
    'please-set-secret-key-in-production'
)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employees.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
csrf = CSRFProtect(app)

# ──────────────────────────────────────────────
#  自动创建数据库表（gunicorn/production 入口）
# ──────────────────────────────────────────────
with app.app_context():
    db.create_all()
    # 创建默认管理员（仅当没有用户时）
    if not User.query.first():
        admin = User(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print('默认管理员已创建：admin / admin123 请及时修改密码！')


# ──────────────────────────────────────────────
#  登录认证
# ──────────────────────────────────────────────

from functools import wraps


def login_required(f):
    """登录校验装饰器：未登录用户重定向到登录页"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    # 已登录则跳转到首页
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('❌ 请输入用户名和密码', 'danger')
            return render_template('login.html')

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            flash(f'✅ 欢迎回来，{user.username}！', 'success')
            return redirect(url_for('index'))
        else:
            flash('❌ 用户名或密码错误', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """登出"""
    session.clear()
    flash('✅ 已安全退出', 'success')
    return redirect(url_for('login'))


# ══════════════════════════════════════════════
#  Web 页面路由
# ══════════════════════════════════════════════

@app.route('/')
@login_required
def index() -> str:
    """员工列表首页，支持搜索和筛选"""
    search_query: str = request.args.get('q', '').strip()
    gender_filter: str = request.args.get('gender', '').strip()
    dept_filter: str = request.args.get('department', '').strip()
    page: int = request.args.get('page', 1, type=int)
    per_page: int = 20

    query = Employee.query

    if search_query:
        query = query.filter(
            Employee.name.contains(search_query) |
            Employee.team_station.contains(search_query) |
            Employee.position.contains(search_query) |
            Employee.department.contains(search_query)
        )

    if gender_filter:
        query = query.filter(Employee.gender == gender_filter)

    if dept_filter:
        query = query.filter(Employee.department == dept_filter)

    pagination = query.order_by(Employee.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    employees = pagination.items

    # 获取所有部门（用于筛选下拉框）
    departments = db.session.query(Employee.department).distinct().all()
    departments = sorted([d[0] for d in departments if d[0]])

    return render_template(
        'index.html',
        employees=employees,
        pagination=pagination,
        search_query=search_query,
        gender_filter=gender_filter,
        dept_filter=dept_filter,
        departments=departments
    )


@app.route('/add', methods=['GET', 'POST'])
@login_required
def add() -> str:
    """新增员工"""
    if request.method == 'POST':
        valid: bool
        errors: dict
        valid, errors = _validate_employee_data(request.form)
        if not valid:
            for field, msg in errors.items():
                flash(f'❌ {msg}', 'danger')
            return render_template('add.html')

        try:
            hire_date, date_err = _parse_date(request.form.get('hire_date'))
            if date_err:
                flash(f'❌ {date_err}', 'danger')
                return render_template('add.html')

            employee = Employee(
                name=request.form['name'].strip(),
                gender=request.form['gender'].strip(),
                age=_parse_int(request.form.get('age')),
                birthplace=request.form.get('birthplace', '').strip() or None,
                department=request.form.get('department', '').strip() or None,
                position=request.form.get('position', '').strip() or None,
                team_station=request.form.get('team_station', '').strip() or None,
                work_content=request.form.get('work_content', '').strip() or None,
                remarks=request.form.get('remarks', '').strip() or None,
                phone=request.form.get('phone', '').strip() or None,
                email=request.form.get('email', '').strip() or None,
                hire_date=hire_date,
            )
            db.session.add(employee)
            db.session.commit()
            flash('✅ 员工添加成功！', 'success')
            return redirect(url_for('index'))
        except (ValueError, SQLAlchemyError) as e:
            db.session.rollback()
            flash(f'❌ 添加失败：{str(e)}', 'danger')

    return render_template('add.html')


@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id: int) -> str:
    """编辑员工信息"""
    employee = Employee.query.get_or_404(id)

    if request.method == 'POST':
        valid, errors = _validate_employee_data(request.form)
        if not valid:
            for field, msg in errors.items():
                flash(f'❌ {msg}', 'danger')
            return render_template('edit.html', employee=employee)

        try:
            employee.name = request.form['name'].strip()
            employee.gender = request.form['gender'].strip()
            employee.age = _parse_int(request.form.get('age'))
            employee.birthplace = request.form.get('birthplace', '').strip() or None
            employee.department = request.form.get('department', '').strip() or None
            employee.position = request.form.get('position', '').strip() or None
            employee.team_station = request.form.get('team_station', '').strip() or None
            employee.work_content = request.form.get('work_content', '').strip() or None
            employee.remarks = request.form.get('remarks', '').strip() or None
            employee.phone = request.form.get('phone', '').strip() or None
            employee.email = request.form.get('email', '').strip() or None
            hire_date, date_err = _parse_date(request.form.get('hire_date'))
            if date_err:
                flash(f'❌ {date_err}', 'danger')
                return render_template('edit.html', employee=employee)
            employee.hire_date = hire_date
            db.session.commit()
            flash('✅ 员工信息已更新！', 'success')
            return redirect(url_for('index'))
        except (ValueError, SQLAlchemyError) as e:
            db.session.rollback()
            flash(f'❌ 更新失败：{str(e)}', 'danger')

    return render_template('edit.html', employee=employee)


@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id: int) -> str:
    """删除员工"""
    employee = Employee.query.get_or_404(id)
    try:
        db.session.delete(employee)
        db.session.commit()
        flash('✅ 员工已删除！', 'success')
    except (ValueError, SQLAlchemyError) as e:
        flash(f'❌ 删除失败：{str(e)}', 'danger')
    return redirect(url_for('index'))


@app.route('/export/csv')
@login_required
def export_csv() -> Response:
    """导出员工数据为 CSV"""
    employees = Employee.query.order_by(Employee.id).all()
    output = io.StringIO()
    writer = csv.writer(output)

    # 表头
    writer.writerow(['编号', '姓名', '性别', '岗位/职务', '现班组/井站',
                     '分管工作内容', '备注', '年龄', '出生地',
                     '部门', '电话', '邮箱', '入职日期'])

    for emp in employees:
        writer.writerow([
            emp.id, emp.name, emp.gender, emp.position, emp.team_station,
            emp.work_content, emp.remarks, emp.age, emp.birthplace,
            emp.department, emp.phone, emp.email,
            emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else ''
        ])

    output.seek(0)
    return _send_csv_response(output, 'employees.csv')


@app.route('/export/excel')
@login_required
def export_excel() -> Response:
    """导出员工数据为 Excel (.xlsx)"""
    employees = Employee.query.order_by(Employee.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = '员工信息'

    # 表头
    headers = ['编号', '姓名', '性别', '岗位/职务', '现班组/井站',
               '分管工作内容', '备注', '年龄', '出生地',
               '部门', '电话', '邮箱', '入职日期']
    ws.append(headers)

    for emp in employees:
        ws.append([
            emp.id, emp.name, emp.gender, emp.position, emp.team_station,
            emp.work_content, emp.remarks, emp.age, emp.birthplace,
            emp.department, emp.phone, emp.email,
            emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else ''
        ])

    # 调整列宽
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employees.xlsx'
    )


@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_employees():
    """导入 Excel 员工数据"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('❌ 请选择要导入的 Excel 文件', 'danger')
            return render_template('import.html')

        # 判断文件格式
        filename = file.filename.lower()
        if filename.endswith('.xlsx'):
            try:
                wb = load_workbook(file)
            except Exception:
                flash('❌ 无法读取 .xlsx 文件，请检查格式', 'danger')
                return render_template('import.html')
        elif filename.endswith('.xls'):
            try:
                # 用 xlrd 读取旧格式
                import xlrd
                xls_wb = xlrd.open_workbook(file_contents=file.read())
                xls_ws = xls_wb.sheet_by_index(0)
                wb = _xlrd_to_openpyxl(xls_ws)
            except Exception as e:
                flash(f'❌ 无法读取 .xls 文件：{str(e)}', 'danger')
                return render_template('import.html')
        else:
            flash('❌ 仅支持 .xls 和 .xlsx 格式', 'danger')
            return render_template('import.html')

        try:
            ws = wb.active
            header_row = [str(c.value or '').strip() for c in ws[1]]

            # 字段名映射（支持中英文表头）
            field_map = _build_field_map(header_row)

            success_count = 0
            error_count = 0
            errors_detail = []

            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, field_name in field_map.items():
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    row_data[field_name] = str(cell.value).strip() if cell.value is not None else ''

                name = row_data.get('name', '')
                if not name:
                    error_count += 1
                    errors_detail.append(f'第 {row_idx} 行：姓名为空')
                    continue

                try:
                    employee = Employee(
                        name=name,
                        gender=row_data.get('gender', ''),
                        age=_parse_int(row_data.get('age')),
                        birthplace=row_data.get('birthplace') or None,
                        department=row_data.get('department') or None,
                        position=row_data.get('position') or None,
                        team_station=row_data.get('team_station') or None,
                        work_content=row_data.get('work_content') or None,
                        remarks=row_data.get('remarks') or None,
                        phone=row_data.get('phone') or None,
                        email=row_data.get('email') or None,
                        hire_date=_parse_date(row_data.get('hire_date'))[0],
                    )
                    db.session.add(employee)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors_detail.append(f'第 {row_idx} 行（{name}）：{str(e)}')

            db.session.commit()
            msg = f'✅ 导入完成：成功 {success_count} 条'
            if error_count:
                msg += f'，失败 {error_count} 条'
                for err in errors_detail[:5]:
                    msg += f'<br>  - {err}'
                if len(errors_detail) > 5:
                    msg += f'<br>  - ...及其他 {len(errors_detail) - 5} 条错误'
            flash(msg, 'success' if error_count == 0 else 'warning')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'❌ 导入失败：{str(e)}', 'danger')

    return render_template('import.html')


# ══════════════════════════════════════════════
#  RESTful API 路由
# ══════════════════════════════════════════════

@app.route('/api/employees', methods=['GET'])
def api_list() -> Response:
    """API: 获取员工列表（支持搜索）"""
    search = request.args.get('q', '').strip()
    query = Employee.query
    if search:
        query = query.filter(
            Employee.name.contains(search) |
            Employee.birthplace.contains(search) |
            Employee.department.contains(search)
        )
    employees = query.order_by(Employee.id.desc()).all()
    return jsonify({'code': 200, 'data': [e.to_dict() for e in employees]})


@app.route('/api/employees/<int:id>', methods=['GET'])
def api_detail(id: int) -> Response:
    """API: 获取单个员工信息"""
    employee = Employee.query.get_or_404(id)
    return jsonify({'code': 200, 'data': employee.to_dict()})


@app.route('/api/employees', methods=['POST'])
@csrf.exempt
def api_create() -> Response:
    """API: 新增员工"""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'code': 400, 'message': '请求体不能为空'}), 400

    required_fields = ['name', 'gender', 'age', 'birthplace']
    for field in required_fields:
        if field not in data:
            return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400

    # 格式验证
    valid, errors = _validate_employee_data(data)
    if not valid:
        msg = '；'.join(errors.values())
        return jsonify({'code': 400, 'message': msg}), 400

    hire_date, date_err = _parse_date(data.get('hire_date'))
    if date_err and data.get('hire_date'):
        return jsonify({'code': 400, 'message': date_err}), 400

    try:
        employee = Employee(
            name=data['name'],
            gender=data['gender'],
            age=_parse_int(data.get('age')),
            birthplace=data.get('birthplace') or None,
            department=data.get('department') or None,
            position=data.get('position') or None,
            team_station=data.get('team_station') or None,
            work_content=data.get('work_content') or None,
            remarks=data.get('remarks') or None,
            phone=data.get('phone') or None,
            email=data.get('email') or None,
            hire_date=hire_date,
        )
        db.session.add(employee)
        db.session.commit()
        return jsonify({'code': 201, 'message': '创建成功', 'data': employee.to_dict()}), 201
    except (ValueError, SQLAlchemyError) as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@app.route('/api/employees/<int:id>', methods=['PUT'])
@csrf.exempt
def api_update(id: int) -> Response:
    """API: 更新员工信息"""
    employee = Employee.query.get_or_404(id)
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'code': 400, 'message': '请求体不能为空'}), 400

    try:
        # 验证提供的数据（仅检查实际传入的字段）
        single_field_errors: dict[str, str] = {}

        if 'name' in data:
            v = data['name']
            if not v or not str(v).strip():
                single_field_errors['name'] = '姓名不能为空'
            elif len(str(v)) > 50:
                single_field_errors['name'] = '姓名不能超过 50 个字符'
        if 'gender' in data:
            v = data['gender']
            if v not in ('男', '女'):
                single_field_errors['gender'] = '性别只能为「男」或「女」'
        if 'age' in data:
            try:
                age_val = int(data['age'])
                if age_val < 16 or age_val > 100:
                    single_field_errors['age'] = '年龄必须在 16 到 100 之间'
            except (ValueError, TypeError):
                single_field_errors['age'] = '年龄必须是有效数字'
        if 'birthplace' in data:
            v = data['birthplace']
            if not v or not str(v).strip():
                single_field_errors['birthplace'] = '出生地不能为空'
            elif len(str(v)) > 100:
                single_field_errors['birthplace'] = '出生地不能超过 100 个字符'
        if 'phone' in data and data['phone']:
            if not _is_valid_phone(str(data['phone'])):
                single_field_errors['phone'] = '电话格式不正确'
        if 'email' in data and data['email']:
            if not _is_valid_email(str(data['email'])):
                single_field_errors['email'] = '邮箱格式不正确'

        if single_field_errors:
            return jsonify({'code': 400, 'message': '；'.join(single_field_errors.values())}), 400

        if 'name' in data:
            employee.name = data['name']
        if 'gender' in data:
            employee.gender = data['gender']
        if 'age' in data:
            employee.age = _parse_int(data['age'])
        if 'birthplace' in data:
            employee.birthplace = data['birthplace']
        if 'department' in data:
            employee.department = data['department']
        if 'position' in data:
            employee.position = data['position']
        if 'team_station' in data:
            employee.team_station = data['team_station']
        if 'work_content' in data:
            employee.work_content = data['work_content']
        if 'remarks' in data:
            employee.remarks = data['remarks']
        if 'phone' in data:
            employee.phone = data['phone']
        if 'email' in data:
            employee.email = data['email']
        if 'hire_date' in data:
            hire_date, date_err = _parse_date(data['hire_date'])
            if date_err and data.get('hire_date'):
                return jsonify({'code': 400, 'message': date_err}), 400
            employee.hire_date = hire_date

        db.session.commit()
        return jsonify({'code': 200, 'message': '更新成功', 'data': employee.to_dict()})
    except (ValueError, SQLAlchemyError) as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500


@app.route('/api/employees/<int:id>', methods=['DELETE'])
@csrf.exempt
def api_delete(id: int) -> Response:
    """API: 删除员工"""
    employee = Employee.query.get_or_404(id)
    try:
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except SQLAlchemyError as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500


# ──────────────────────────────────────────────
#  数据验证
# ──────────────────────────────────────────────

def _validate_employee_data(data):
    """验证员工表单/API 数据，返回 (is_valid, errors_dict)

    data 是类字典对象（request.form 或 JSON dict）。
    errors 是 {字段名: 错误消息} 的字典。
    """
    errors = {}

    name = data.get('name', '').strip()
    if not name:
        errors['name'] = '姓名不能为空'
    elif len(name) > 50:
        errors['name'] = '姓名不能超过 50 个字符'

    gender = data.get('gender', '').strip()
    if not gender:
        errors['gender'] = '性别不能为空'
    elif gender not in ('男', '女'):
        errors['gender'] = '性别只能为「男」或「女」'

    age_raw = data.get('age', '')
    if age_raw:
        try:
            age = int(age_raw)
            if age < 16 or age > 100:
                errors['age'] = '年龄必须在 16 到 100 之间'
        except (ValueError, TypeError):
            errors['age'] = '年龄必须是有效数字'

    birthplace = data.get('birthplace', '').strip()
    if birthplace and len(birthplace) > 100:
        errors['birthplace'] = '出生地不能超过 100 个字符'

    phone = data.get('phone', '').strip()
    if phone and not _is_valid_phone(phone):
        errors['phone'] = '电话格式不正确（应为 11 位手机号或带区号座机）'

    email = data.get('email', '').strip()
    if email and not _is_valid_email(email):
        errors['email'] = '邮箱格式不正确'

    return len(errors) == 0, errors


def _is_valid_phone(phone: str) -> bool:
    """简单手机号/座机格式校验"""
    # 手机号：11 位数字，以 1 开头
    # 座机：区号-号码，如 010-88888888
    return bool(re.match(r'^1\d{10}$', phone) or
                re.match(r'^0\d{2,3}-?\d{7,8}$', phone))


def _is_valid_email(email: str) -> bool:
    """简单邮箱格式校验"""
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email))


# ──────────────────────────────────────────────
#  工具函数
# ──────────────────────────────────────────────

def _parse_date(date_str: Optional[str]) -> tuple[Optional[date], Optional[str]]:
    """解析日期字符串，返回 (date, error_message) 元组。

    解析成功时返回 (date_obj, None)；失败时返回 (None, '日期格式不正确')。
    """
    if not date_str or not date_str.strip():
        return None, None
    try:
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date(), None
    except ValueError:
        return None, '日期格式不正确（应为 YYYY-MM-DD）'


def _build_field_map(header_row):
    """将 Excel 表头映射到 Employee 模型字段名"""
    mapping = {
        '姓名': 'name', 'name': 'name',
        '性别': 'gender', 'gender': 'gender',
        '年龄': 'age', 'age': 'age',
        '出生地': 'birthplace', 'birthplace': 'birthplace',
        '部门': 'department', 'department': 'department',
        '岗位': 'position', '职务': 'position',
        '岗位/职务': 'position', '职位': 'position', 'position': 'position',
        '现班组': 'team_station', '井站': 'team_station',
        '现班组/井站': 'team_station', '班组': 'team_station', 'team_station': 'team_station',
        '分管工作内容': 'work_content', '工作内容': 'work_content',
        'work_content': 'work_content',
        '备注': 'remarks', 'remark': 'remarks', 'remarks': 'remarks',
        '电话': 'phone', '联系电话': 'phone', 'phone': 'phone',
        '邮箱': 'email', '电子邮箱': 'email', 'email': 'email',
        '入职日期': 'hire_date', '入职时间': 'hire_date', 'hire_date': 'hire_date',
        '序号': None, 'id': None, '编号': None,
    }
    field_map = {}
    for col_idx, header in enumerate(header_row):
        key = header.strip().replace(' ', '')
        if key in mapping and mapping[key] is not None:
            field_map[col_idx] = mapping[key]
        # 中英文分号分隔
        elif '/' in key:
            for part in key.split('/'):
                part = part.strip()
                if part in mapping and mapping[part] is not None:
                    field_map[col_idx] = mapping[part]
                    break
    return field_map


def _xlrd_to_openpyxl(xls_ws):
    """将 xlrd 的 sheet 转换为 openpyxl 的 Workbook"""
    wb = Workbook()
    ws = wb.active
    for row_idx in range(xls_ws.nrows):
        for col_idx in range(xls_ws.ncols):
            cell = xls_ws.cell(row_idx, col_idx)
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
    return wb


def _parse_int(value):
    """安全地解析整数，失败返回 None"""
    if not value:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _send_csv_response(output: io.StringIO, filename: str) -> Response:
    """将 CSV 字符串流作为附件响应返回"""
    output_bytes = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    return Response(
        output_bytes.getvalue(),
        mimetype='text/csv; charset=utf-8-sig',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'text/csv; charset=utf-8-sig'
        }
    )


# ──────────────────────────────────────────────
#  统一错误处理器
# ──────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e) -> tuple[str, int]:
    """404 页面"""
    return render_template('error.html', code=404, message='页面未找到'), 404


@app.errorhandler(500)
def server_error(e) -> tuple[str, int]:
    """500 页面"""
    return render_template('error.html', code=500, message='服务器内部错误'), 500


# ──────────────────────────────────────────────
#  CLI 命令
# ──────────────────────────────────────────────

@app.cli.command('init-db')
def init_db() -> None:
    """初始化数据库（创建所有表）"""
    db.create_all()
    print('数据库表已创建。')


# ──────────────────────────────────────────────
#  启动入口
# ──────────────────────────────────────────────

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)
